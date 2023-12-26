"""
By: Aaron J. | Contains the functions open_onedrive, save_prefabs, get_prefabs, get_time and Device class.
"""
from openpyxl import load_workbook
from resources.temperature import Temperature_Probe
from time import sleep
import datetime
import json

def open_onedrive(type_run):
    print("Starting OneDrive...")
    if type_run:
        import os
        try:
            print("OneDrive Mounted: TYPE-A")
            os.system("rclone --vfs-cache-mode writes mount onedrive: /home/pi/Desktop/onedrive/")
        except Exception as exc:
            print(f"Error: {exc}")
    else:
        print("OneDrive Mounted: TYPE-F")


def save_prefabs(data):
    with open("resources/data/prefabs.json", "w") as json_file:
        json.dump(data, json_file)


def get_prefabs():
    try:
        with open("resources/data/prefabs.json", "r") as json_file:
            data = json.load(json_file)
            return data
    except FileNotFoundError:
        print("File not found or doesn't exist yet.")
        return None


def get_time():
    now = datetime.datetime.now()
    today = datetime.date.today()
    date = today.strftime("%B %d %Y")
    time = "AM"
    current_hour = int(now.strftime("%H"))
    current_minutes = str(now.strftime("%M"))
    if current_hour > 12:
        current_hour -= 12
        time = "PM"
    if current_hour == 0:
        current_hour = 12
    current_time = f"{current_hour}:{current_minutes} {time}"
    return current_time


class Device:
    def __init__(self, name, location):
        self.name = name
        self.location = location

    def start_device(self):
        prefabs = get_prefabs()

        sleep(10)  # Waiting for OneDrive
        device_location = "/home/pi/Desktop/onedrive/My_Temperatures/SITE-1/Temperature_Log.xlsx"
        wb = load_workbook(device_location)
        ws = wb.active
        probe = Temperature_Probe("Default-Name", 1)

        if not prefabs["CONFIG"]["HAS_RUN"]:
            # Header
            ws["C1"].value = f"{self.name}"
            ws["G1"].value = f"{self.location}"
            ws["A1"].value = "Remotely Cool"
            ws["A2"].value = "########################################################################"

            # Meat and potatoes
            ws["A3"].value = "Time"
            ws["C3"].value = "Temperature"
            ws["E3"].value = "Notes"
            ws["G3"].value = "Date"
            prefabs["CONFIG"]["HAS_RUN"] = True
            save_prefabs(prefabs)
            print("First Time Setup Complete.")

        # Getting real data
        temp_f, temp_c = probe.action()
        note = None
        temp_f = float(temp_f)
        previous_temp_f = temp_f
        differential = 3  # Amount of change before temp alert
        warn_before = 5  # Amount before max or min temp alert
        if temp_f > previous_temp_f + differential or temp_f < previous_temp_f - differential:
            note = "SHARP CHANGE"
        if temp_f > prefabs["CONFIG"]["MAX_TEMP"]:
            note = "TEMP TOO HIGH"
        elif temp_f < prefabs["CONFIG"]["MIN_TEMP"]:
            note = "TEMP TOO LOW"
        elif temp_f > prefabs["CONFIG"]["MAX_TEMP"] - warn_before:
            note = "SLIGHTLY HIGH"
        elif temp_f < prefabs["CONFIG"]["MIN_TEMP"] + warn_before:
            note = "SLIGHTLY LOW"

        ws.insert_rows(4)
        ws["A4"].value = get_time()
        ws["C4"].value = float(probe.action()[0])
        ws["E4"].value = note
        ws["G4"].value = datetime.date.today()
        # for headers
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=2)
        ws.merge_cells(start_row=1, end_row=1, start_column=3, end_column=4)
        ws.merge_cells(start_row=1, end_row=1, start_column=5, end_column=6)
        ws.merge_cells(start_row=1, end_row=1, start_column=7, end_column=8)
        # For data
        ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=8)
        ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=2)
        ws.merge_cells(start_row=3, end_row=3, start_column=3, end_column=4)
        ws.merge_cells(start_row=3, end_row=3, start_column=5, end_column=6)
        ws.merge_cells(start_row=3, end_row=3, start_column=7, end_column=8)
        for x in range(100):
            # ws.merge_cells(start_row=4 + x, end_row=4 + x, start_column=1, end_column=2)  # For Time Column
            # ws.merge_cells(start_row=4 + x, end_row=4 + x, start_column=3, end_column=4)  # For Temperature Column
            ws.merge_cells(start_row=4 + x, end_row=4 + x, start_column=5, end_column=6)  # For Notes Column
            ws.merge_cells(start_row=4 + x, end_row=4 + x, start_column=7, end_column=8)  # For Date Column

        while True:
            try:
                wb.save(device_location)
                print(f"[{get_time()}] Temperature: {temp_f}F° For '{self.name}' Saved.")
                break
            except PermissionError:
                print("Permission Denied. Close Any Open Excel Files.")
            except Exception as exc:
                print(f"Error E: {exc}")

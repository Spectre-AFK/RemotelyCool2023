from openpyxl import load_workbook, Workbook
import pickle, datetime
from datetime import datetime
from resources.temperature import Temperature_Probe
import threading
from time import sleep

def open_onedrive(type_run):
    if type_run:
        import os
        try:
            print("OneDrive Mounting: TYPE-A")
            os.system("rclone --vfs-cache-mode writes mount onedrive: /home/pi/Desktop/onedrive/")
        except Exception as exc:
            print(f"Error: {exc}")
    else:
        print("OneDrive Mounted: TYPE-F")

def get_time():
    from datetime import date
    now = datetime.now()
    today = date.today()
    date = today.strftime("%B %d %Y")
    time = "AM"
    current_hour = int(now.strftime("%H"))
    current_minutes = str(now.strftime("%M"))
    if current_hour > 12:
        current_hour -= 12
        time = "PM"
    if current_hour == 0:
        current_hour = 12
    current_time = f"{current_hour}:{current_minutes} {time}"
    return current_time

def save_file(name, content):
    file = f"resources/data/{name}.dat"
    with open(file, "wb") as f:
        pickle.dump(content, f)
        print(f"File '{name}' has been saved.")

def find_file(name):
    file = f"resources/data/{name}.dat"
    try:
        with open(file, "rb") as f:
            result = pickle.load(f)
            return result
    except Exception:
        return False

def has_run():
    run = find_file("run")
    if run == False:
        save_file("run", "True")
        return False
    else:
        return True

def confirm(text):  # Confirm some prompt
    while True:
        temp = input(text)
        try:
            result = input(f"Confirm: '{temp}'\n(Y/N): ").lower()
        except Exception:
            print("Invalid Input")
        if result == "y":
            return temp


class Device:
    def __init__(self, name, location):
        self.name = name
        self.location = location

    def start_device(self):
        from datetime import date
        device_location = "/home/pi/Desktop/onedrive/My_Temperatures/ONSITE/Temperature_Log.xlsx"
        print("Awaiting OneDrive...")
        sleep(5)
        wb = load_workbook(device_location)
        ws = wb.active
        probe = Temperature_Probe("Default-Name", 1)

        if has_run() == False:
            ws["C1"].value = f"{self.name}"
            ws["G1"].value = f"{self.location}"
            ws["A1"].value = "Remotely Cool"

            ws["A3"].value = "Time"
            ws["C3"].value = "Temperature"
            ws["E3"].value = "Notes"
            ws["G3"].value = "Date"
            print("First Time Setup Complete.")

        ws.insert_rows(4)
        ws["A4"].value = get_time()
        ws["C4"].value = float(probe.action()[0])
        ws["E4"].value = "N/A" 
        ws["G4"].value = date.today()       
        while True:
            try:
                wb.save(device_location)
                print(f"[{get_time()}] Temperature: {float(probe.action()[0])}F° For '{self.name}' Saved.")
                break
            except PermissionError:
                print("Permission Denied. Close Any Open Excel Files.")
            except Exception as exc:
                print(f"Error E: {exc}")